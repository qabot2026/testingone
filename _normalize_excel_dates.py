"""Normalize Excel Conv. Date / Conv. Time to matching text (separate columns).

Optional: add --sheet2 to refresh a dashboard tab (same KPIs as the web viewer)
via excel_sheet2_analysis.py on the saved workbook.
"""
from __future__ import annotations

import argparse
from datetime import date, datetime, time
from pathlib import Path

import openpyxl

XLSX = Path(__file__).resolve().parent / "Excel.xlsx"
OUT = Path(__file__).resolve().parent / "Excel_datetime_formatted.xlsx"


def sheet_date_str(d: date) -> str:
    return f"{d.day} {d.strftime('%b %Y')}"


def sheet_time_str(t: time) -> str:
    hour = t.hour
    h12 = hour % 12
    if h12 == 0:
        h12 = 12
    ampm = "AM" if hour < 12 else "PM"
    return f"{h12}:{t.minute:02d}:{t.second:02d} {ampm}"


def parse_row_date_time(a_val, b_val) -> tuple[date, time] | None:
    if a_val is None:
        return None
    if isinstance(a_val, datetime):
        d = a_val.date()
        if isinstance(b_val, time):
            return d, b_val
        tt = a_val.time()
        if tt != time(0, 0, 0):
            return d, tt
        if isinstance(b_val, time):
            return d, b_val
        return d, time(0, 0, 0)
    if isinstance(a_val, date):
        d = a_val
        if isinstance(b_val, time):
            return d, b_val
        return d, time(0, 0, 0)
    return None


def main() -> None:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument(
        "--sheet2",
        action="store_true",
        help="After save, run excel_sheet2_analysis on the output file (adds/updates Sheet2).",
    )
    args = p.parse_args()

    wb = openpyxl.load_workbook(XLSX)
    ws = wb.active
    n = 0
    for row_idx in range(2, ws.max_row + 1):
        cell_a = ws.cell(row_idx, 1)
        cell_b = ws.cell(row_idx, 2)
        parsed = parse_row_date_time(cell_a.value, cell_b.value)
        if not parsed:
            continue
        d, tm = parsed
        cell_a.value = sheet_date_str(d)
        cell_b.value = sheet_time_str(tm)
        cell_a.number_format = "@"
        cell_b.number_format = "@"
        n += 1
    try:
        wb.save(XLSX)
        out_path = XLSX
    except PermissionError:
        wb.save(OUT)
        out_path = OUT
        print("NOTE: Excel.xlsx is open or locked. Wrote sidecar file instead.")
    print(f"Normalized {n} rows -> {out_path}")
    if args.sheet2:
        try:
            from excel_sheet2_analysis import write_analysis_workbook

            write_analysis_workbook(out_path, source_sheet="Sheet1", analysis_sheet="Sheet2")
            print("Sheet2 analysis refreshed.")
        except Exception as e:
            print(f"Sheet2 step skipped: {e}")


if __name__ == "__main__":
    main()
