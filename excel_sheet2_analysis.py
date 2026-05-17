"""
Build Sheet2 in the same workbook with the same KPIs as the conversations-sheet
web dashboard (totals, channels, contact segments, per-segment channel split).

Reads row data from Sheet1 (or --source), matching columns by header names compatible
with the chatbot Google Sheet (mobile, email, channel, appointment*, conv date, etc.).
Requires: pip install openpyxl
"""
from __future__ import annotations

import argparse
import re
from dataclasses import dataclass, field
from datetime import date, datetime, time
from pathlib import Path

import openpyxl
from openpyxl.styles import Font

# --- Header aliases (aligned with server/chatbot-api/lib/sheets.mjs) ---

CONV_DATE_ALIASES = [
    "conversationdate",
    "convdate",
    "convdateonly",
    "conversiondate",
    "date",
]
MOBILE_ALIASES = [
    "mobile",
    "phone",
    "phonenumber",
    "mobilenumber",
    "mobile_number",
    "tel",
    "cell",
    "cellphone",
    "contactnumber",
    "contactphone",
    "whatsappnumber",
    "yourmobile",
    "usermobile",
    "customermobile",
    "mobile_no",
    "mobileno",
    "phone_no",
    "phoneno",
    "cell_number",
]
EMAIL_ALIASES = [
    "email",
    "mail",
    "e_mail",
    "email_address",
    "emailaddress",
    "useremail",
    "contactemail",
    "contact_email",
    "email_id",
    "e_mail_address",
    "mail_id",
]
CHANNEL_ALIASES = [
    "channel",
    "channels",
    "whatsapp",
    "whatsappchannel",
    "chatsource",
    "sourcechannel",
    "communicationchannel",
    "chatchannel",
    "userchannel",
    "entrychannel",
    "originchannel",
    "platformchannel",
]
APPT_BOOKED_ALIASES = [
    "appointmentbooked",
    "appointment_booked",
    "isappointmentbooked",
    "appointmentscheduled",
    "appointmentstatus",
    "appointmentbookingstatus",
    "appointmentconfirmation",
    "consultbooked",
    "consultscheduled",
    "appointmentdone",
    "apptstatus",
    "apptbooking",
    "apptscheduled",
    "consultbooking",
    "bookedappointment",
    "bookingdone",
    "bookingstatus",
]
APPT_DATE_ALIASES = [
    "appointmentdate",
    "apptdate",
    "appointmentday",
    "selectedappointmentdate",
    "appointmentpickeddate",
    "dateofappointment",
    "scheduleddate",
    "apptscheduleddate",
]
APPT_TIME_ALIASES = [
    "appointmenttime",
    "appttime",
    "appointmenttimeslot",
    "scheduledtime",
    "apptscheduledtime",
    "slottime",
]
APPT_DATETIME_ALIASES = [
    "appointmentdatetime",
    "appointment_date_time",
    "apptdatetime",
    "appointment_at",
    "scheduledatetime",
    "scheduledat",
    "bookedat",
    "booked_datetime",
]

REX_WEB = re.compile(
    r"\bwebsite\b|^web([\s_-]|$)|^web$|\bwebchat\b|\bwebview\b|^inappwebview|"
    r"(^|[\s,])www\.|\bbrowser\b|(^|[\s,])desktop\b|\bportal\b|^online([\s_-]|$)|"
    r"^online$|\binternet\b|^internet$|^site$|^www$|^cx\b|^sse\b|^widget\b|"
    r"^embed(ded)?\b|^hosted\b|^organic\b|^direct\b",
    re.I,
)
REX_WHATSAPP = re.compile(
    r"\bwhatsapp\b|(^|[\s,])wa([\s,/]|$)|whats[\s_-]*app",
    re.I,
)
REX_INSTAGRAM = re.compile(r"\binstagram\b|(^|[\s,])ig([\s,/]|$)", re.I)
REX_FACEBOOK = re.compile(
    r"\bfacebook\b|(^|[\s,])fb([\s,/]|$)|\bmessenger\b|meta[\s_-]*business",
    re.I,
)
REX_EMAIL_OK = re.compile(r"^[^\s@]+@[^\s@]+\.[^\s@]+")
REX_APPT_POSITIVE = re.compile(
    r"\b(scheduled|booked)\b|appointment\s*(booked|scheduled|confirmed|fixed|set)|"
    r"booking\s*(done|confirmed|complete)|\bconfirmation\b|\bconfirmed\b|"
    r"consult(ation)?\s*(booked|scheduled)|slot\s*(reserved|booked)|^completed$",
    re.I,
)
REX_APPT_DTTM_COMBO = re.compile(
    r"\d{4}-\d{2}-\d{2}[T ]\d{1,2}:\d{2}|\d{4}-\d{2}-\d{2}[T ]\d{1,2}(?!:)|"
    r"\d{1,4}[/.\-]\d{1,4}[/.\-]\d{2,4}.*\d{1,2}:\d{2}",
)


def norm_header_key(val) -> str:
    return re.sub(r"[^a-z0-9]", "", str(val or "").strip().lower())


def build_header_index_row(ws) -> dict[str, int]:
    """Normalized header -> 1-based column index (first occurrence wins)."""
    out: dict[str, int] = {}
    for col in range(1, (ws.max_column or 0) + 1):
        raw = ws.cell(1, col).value
        k = norm_header_key(raw)
        if k and k not in out:
            out[k] = col
    return out


def pick_column(header_index: dict[str, int], aliases: list[str], fallback_1based: int) -> int:
    for a in aliases:
        nk = norm_header_key(a)
        if nk in header_index:
            return header_index[nk]
    return fallback_1based


def first_datetime_alias_col(header_index: dict[str, int], aliases: list[str]) -> int | None:
    for a in aliases:
        nk = norm_header_key(a)
        if nk in header_index:
            return header_index[nk]
    return None


def cell_str(raw) -> str:
    if raw is None:
        return ""
    if isinstance(raw, datetime):
        return raw.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(raw, date):
        return raw.isoformat()
    if isinstance(raw, time):
        return raw.strftime("%H:%M:%S")
    return str(raw).strip()


def row_nonempty(values: tuple) -> bool:
    for v in values:
        if v is None:
            continue
        if isinstance(v, str) and not v.strip():
            continue
        return True
    return False


def has_lead_mobile(raw) -> bool:
    if has_lead_email(raw):
        return False
    digits = re.sub(r"\D", "", cell_str(raw))
    return len(digits) >= 7


def has_lead_email(raw) -> bool:
    t = cell_str(raw)
    if not t or "@" not in t:
        return False
    return bool(REX_EMAIL_OK.match(t))


def appointment_booked_sheet_value(raw) -> str:
    s = cell_str(raw)
    if not s or re.match(r"^no$", s, re.I):
        return ""
    if re.match(r"^yes$", s, re.I) or re.match(r"^scheduled$", s, re.I):
        return "Scheduled"
    return s


def appointment_cell_counts_scheduled(raw) -> bool:
    cell = cell_str(raw)
    if not cell:
        return False
    if re.match(
        r"^no$|^none$|^false$|^not\s*(scheduled|booked|yet)$|^n/?a$|^pending$|^0$",
        cell,
        re.I,
    ) or re.match(r"^cancel|^reject", cell, re.I):
        return False
    if appointment_booked_sheet_value(raw) == "Scheduled":
        return True
    if re.match(r"^yes$|^true$|^done$|^y$|^1$|^[✓✔☑]", cell, re.I):
        return True
    if REX_APPT_POSITIVE.search(cell):
        return True
    return False


def row_appt_slot_likely_filled(date_raw, time_raw) -> bool:
    d = cell_str(date_raw).strip()
    t = cell_str(time_raw).strip()
    if not d or not t or not re.search(r"\d", d) or not re.search(r"\d", t):
        return False
    if re.search(r"[:.]", t) or re.search(r"\b(am|pm)\b", t, re.I):
        return True
    if re.match(r"^\d{3,4}$", t):
        return True
    if re.search(r"\d{1,2}\s*h", t, re.I):
        return True
    return False


def cell_looks_like_appt_datetime_combined(raw) -> bool:
    s = cell_str(raw).strip()
    if not s or not re.search(r"\d", s):
        return False
    return bool(REX_APPT_DTTM_COMBO.search(s))


def conversation_channel_bucket(raw) -> str:
    s = cell_str(raw).lower()
    if not s:
        return "other"
    if REX_WHATSAPP.search(s):
        return "whatsapp"
    if REX_INSTAGRAM.search(s):
        return "instagram"
    if REX_FACEBOOK.search(s):
        return "facebook"
    if REX_WEB.search(s):
        return "web"
    return "other"


def empty_seg() -> dict[str, int]:
    return {"web": 0, "whatsapp": 0, "instagram": 0, "facebook": 0, "other": 0}


def seg_add(acc: dict[str, int], ch: str) -> None:
    if ch in ("web", "whatsapp", "instagram", "facebook"):
        acc[ch] += 1
    else:
        acc["other"] += 1


@dataclass
class LeadStats:
    conversations: int = 0
    only_mobile: int = 0
    only_email: int = 0
    mobile_and_email: int = 0
    neither: int = 0
    appointment_booked: int = 0
    channel_web: int = 0
    channel_whatsapp: int = 0
    channel_instagram: int = 0
    channel_facebook: int = 0
    channel_other: int = 0
    only_mobile_by_ch: dict = field(default_factory=empty_seg)
    only_email_by_ch: dict = field(default_factory=empty_seg)
    both_by_ch: dict = field(default_factory=empty_seg)


def pick_appt_booked_col(header_index: dict[str, int], max_col: int) -> int:
    for a in APPT_BOOKED_ALIASES:
        nk = norm_header_key(a)
        if nk in header_index:
            return header_index[nk]
    deny = {
        "appointmentdate",
        "appointmenttime",
        "appointmentdatetime",
        "appttime",
        "apptdate",
        "appointmentday",
        "dayofappointment",
        "selectedappointmentdate",
        "appointmentpickeddate",
        "dateofappointment",
    }
    for nk, col in sorted(header_index.items(), key=lambda x: x[1]):
        if not nk or nk in deny:
            continue
        if "appointment" in nk or "booking" in nk or "booked" in nk or "appt" in nk:
            return col
    return min(15, max(1, max_col))


def compute_stats(ws) -> LeadStats:
    hi = build_header_index_row(ws)
    max_col = max(hi.values()) if hi else (ws.max_column or 1)

    date_c = pick_column(hi, CONV_DATE_ALIASES, 1)
    mobile_c = pick_column(hi, MOBILE_ALIASES, 4)
    email_c = pick_column(hi, EMAIL_ALIASES, 5)
    channel_c = pick_column(hi, CHANNEL_ALIASES, 6)
    appt_booked_c = pick_appt_booked_col(hi, max_col)
    appt_date_c = pick_column(hi, APPT_DATE_ALIASES, 16)
    appt_time_c = pick_column(hi, APPT_TIME_ALIASES, 17)
    appt_dt_c = first_datetime_alias_col(hi, APPT_DATETIME_ALIASES)

    st = LeadStats()

    for row in ws.iter_rows(
        min_row=2,
        max_row=ws.max_row or 2,
        min_col=1,
        max_col=max_col,
        values_only=True,
    ):
        if not row_nonempty(row):
            continue

        def gc(c: int):
            i = c - 1
            return row[i] if 0 <= i < len(row) else None

        has_m = has_lead_mobile(gc(mobile_c))
        has_e = has_lead_email(gc(email_c))
        ch = conversation_channel_bucket(gc(channel_c))

        st.conversations += 1
        if has_m and has_e:
            st.mobile_and_email += 1
            seg_add(st.both_by_ch, ch)
        elif has_m:
            st.only_mobile += 1
            seg_add(st.only_mobile_by_ch, ch)
        elif has_e:
            st.only_email += 1
            seg_add(st.only_email_by_ch, ch)
        else:
            st.neither += 1

        appt = appointment_cell_counts_scheduled(gc(appt_booked_c))
        if not appt:
            appt = row_appt_slot_likely_filled(gc(appt_date_c), gc(appt_time_c))
        if not appt and appt_dt_c is not None:
            appt = cell_looks_like_appt_datetime_combined(gc(appt_dt_c))
        if appt:
            st.appointment_booked += 1

        if ch == "web":
            st.channel_web += 1
        elif ch == "whatsapp":
            st.channel_whatsapp += 1
        elif ch == "instagram":
            st.channel_instagram += 1
        elif ch == "facebook":
            st.channel_facebook += 1
        else:
            st.channel_other += 1

    return st


def replace_or_create_sheet(wb, name: str):
    if name in wb.sheetnames:
        idx = wb.sheetnames.index(name)
        wb.remove(wb[name])
        return wb.create_sheet(name, index=min(idx, len(wb.sheetnames)))
    return wb.create_sheet(name)


def write_sheet2(ws, st: LeadStats, generated_at: str) -> None:
    bold = Font(bold=True)
    r = 1
    ws.cell(r, 1, "Lead dashboard (same logic as web viewer)")
    ws.cell(r, 2, generated_at)
    r += 2

    leads_cap = st.only_mobile + st.only_email + st.mobile_and_email
    pct = round(leads_cap * 10000 / st.conversations) / 100 if st.conversations else None

    ws.cell(r, 1, "Total conversations").font = bold
    ws.cell(r, 2, st.conversations)
    r += 1
    ws.cell(r, 1, "Lead capture %").font = bold
    ws.cell(r, 2, pct if pct is not None else "")
    r += 1
    ws.cell(r, 1, "Appointments booked").font = bold
    ws.cell(r, 2, st.appointment_booked)
    r += 2

    ws.cell(r, 1, "Conversations by channel").font = bold
    r += 1
    for label, val in (
        ("Web", st.channel_web),
        ("WhatsApp", st.channel_whatsapp),
        ("Instagram", st.channel_instagram),
        ("Facebook", st.channel_facebook),
        ("Other / uncategorized", st.channel_other),
    ):
        ws.cell(r, 1, label)
        ws.cell(r, 2, val)
        r += 1
    r += 1

    ws.cell(r, 1, "Contact detail captured").font = bold
    r += 1
    headers = (
        "Segment",
        "Total",
        "Web",
        "WhatsApp",
        "Instagram",
        "Facebook",
        "Other",
    )
    for c, h in enumerate(headers, 1):
        cell = ws.cell(r, c, h)
        cell.font = bold
    r += 1

    def seg_row(title: str, total: int, seg: dict) -> None:
        nonlocal r
        ws.cell(r, 1, title)
        ws.cell(r, 2, total)
        ws.cell(r, 3, seg["web"])
        ws.cell(r, 4, seg["whatsapp"])
        ws.cell(r, 5, seg["instagram"])
        ws.cell(r, 6, seg["facebook"])
        ws.cell(r, 7, seg["other"])
        r += 1

    seg_row("Mobile only", st.only_mobile, st.only_mobile_by_ch)
    seg_row("Email only", st.only_email, st.only_email_by_ch)
    seg_row("Mobile & email", st.mobile_and_email, st.both_by_ch)
    r += 1

    ws.cell(r, 1, "Rows with neither mobile nor email (conversation still counted)").font = bold
    ws.cell(r, 2, st.neither)
    r += 1

    # Widen a bit for readability
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 14
    for col in "CDEFG":
        ws.column_dimensions[col].width = 12


def write_analysis_workbook(
    path: Path,
    source_sheet: str | None = None,
    analysis_sheet: str = "Sheet2",
) -> Path:
    wb = openpyxl.load_workbook(path)
    if source_sheet:
        if source_sheet not in wb.sheetnames:
            raise ValueError(f"Sheet {source_sheet!r} not in workbook. Available: {wb.sheetnames}")
        ws_src = wb[source_sheet]
    else:
        if "Sheet1" in wb.sheetnames:
            ws_src = wb["Sheet1"]
        else:
            ws_src = wb[wb.sheetnames[0]]

    st = compute_stats(ws_src)
    ws2 = replace_or_create_sheet(wb, analysis_sheet)
    generated = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    write_sheet2(ws2, st, generated)

    try:
        wb.save(path)
        return path
    except PermissionError:
        out = path.with_name(path.stem + "_with_Sheet2" + path.suffix)
        wb.save(out)
        return out


def main() -> None:
    ap = argparse.ArgumentParser(description="Add Sheet2 analysis tab to Excel workbook.")
    ap.add_argument(
        "xlsx",
        nargs="?",
        default=str(Path(__file__).resolve().parent / "Excel.xlsx"),
        help="Path to .xlsx (default: Excel.xlsx next to this script)",
    )
    ap.add_argument("--source", default="Sheet1", help="Worksheet with lead rows (default Sheet1)")
    ap.add_argument("--analysis-sheet", default="Sheet2", help="Name for the analysis tab")
    args = ap.parse_args()
    path = Path(args.xlsx)
    if not path.is_file():
        raise SystemExit(f"Not found: {path}")
    out = write_analysis_workbook(path, source_sheet=args.source, analysis_sheet=args.analysis_sheet)
    if out != path:
        print(f"NOTE: Original file locked — wrote: {out}")
    else:
        print(f"Updated: {path} (tab {args.analysis_sheet!r})")


if __name__ == "__main__":
    main()
